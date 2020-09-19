"""Turnos

    - turnos_semana(fichero, agente)
        Devuelve una lista de cadenas de texto con los turnos de trabajo
        recogidos en la hoja de Excel.

        Forma de uso:
        >>> turnos_semana('GSEMANAL20.xlsx', 2076)
        ['1F', '1F ', 'FGL', '6F', '2F', 'D', 'D']
"""

from os import rename
from openpyxl import load_workbook


hoja = 'G_BENIDORM'

def turnos_semana(fichero, agente):
    """Abre el fichero Excel y extrae los turnos de trabajo de la hoja
    correspondiente. Devuelve una lista con 7 cadenas de texto correspondientes
    a los turnos de cada día de la semana.
    """

    try:
        # Accedo al fichero
        wb = load_workbook(fichero)
        # Accedo a la hoja del gráfico de Benidorm
        ws = wb[hoja]
    except KeyError:
        print(f"ERROR: No se encuentra la hoja {hoja} en el fichero {fichero}")
        exit(1)
    except FileNotFoundError:
        print(f"ERROR: No se ha encontrado el fichero {fichero}")
        exit(1)

    # Filas en las que se encuentra el CF
    celdas_cf = ws['B10':'B100']
    # Columnas en las que se encuentran los turnos
    col_turnos = range(4, 11)
    # Aquí iré añadiendo mis turnos
    mis_turnos = []

    # Busco la fila que contiene mi CF
    mi_fila = 0
    for fila in celdas_cf:
        for celda in fila:
            if celda.value == agente:
                mi_fila = celda.row

    # No se encuentra el agente
    if mi_fila == 0:
        print(f"ERROR: No se ha encontrado al agente {agente}")
        return []

    # A partir de mi_fila, añado los turnos a mis_turnos
    for columna in col_turnos:
        mis_turnos.append(
            str(ws.cell(row=mi_fila, column=columna).value).strip())

    return mis_turnos


def fecha_grafico(fichero):
    try:
        # Accedo al fichero excel
        wb = load_workbook(fichero)
        # Accedo a la hoja de la fecha
        ws = wb['FECHA']
    except KeyError:
        print(
            f"ERROR: No se encuentra la hoja 'FECHA' en el fichero {fichero}")
        exit(1)
    except FileNotFoundError:
        print(f"ERROR: No se ha encontrado el fichero {fichero}")
        exit(1)

    return ws['C3'].value


def str_doc_date(fichero):
    doc_date = fecha_grafico(fichero)
    return doc_date.strftime('%Y-%m-%d')


def rename_doc(fichero):
    new_name = str_doc_date(fichero) + "_GSEM.xlsx"
    rename(fichero, new_name)
    return new_name
